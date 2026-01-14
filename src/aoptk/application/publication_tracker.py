# ruff: noqa: PLR0913
from __future__ import annotations
import sys
import time
import click
import pandas as pd
from Bio import Entrez
from openpyxl import load_workbook
from aoptk.literature.databases.europepmc import EuropePMC
from aoptk.literature.databases.pubmed import PubMed


@click.command()
@click.option(
    "--read_publications_database_path",
    type=str,
    required=True,
    help="Provide path to the database of read publications",
)
@click.option("--master_table_path", type=str, required=True, help="Provide path to the master table")
@click.option("--search_code", type=str, required=True, help="Provide search code to track this search")
@click.option("--email", type=str, required=False, help="Email address to follow PubMed - NCBI guidelines")
@click.option(
    "--query",
    type=str,
    required=True,
    help="Search term for PubMed or Europe PMC",
)
@click.option(
    "--literature_database",
    type=click.Choice(["pubmed", "europepmc"]),
    required=True,
    help="Database to search: PubMed or Europe PMC",
)
def cli(
    read_publications_database_path: str,
    master_table_path: str,
    email: str,
    search_code: str,
    query: str,
    literature_database: str,
) -> None:
    """Generate publications to read and update master table search codes."""
    database_with_ids = generate_database_with_ids(query, literature_database, email)

    metadata = database_with_ids.get_publications_metadata()

    metadata_df = convert_metadata_structures_to_df(
        search_code,
        query,
        literature_database,
        metadata,
    )

    generate_publications_to_read(read_publications_database_path, search_code, metadata_df)
    update_master_table_search_codes(master_table_path, search_code, metadata_df)


def generate_database_with_ids(query: str, literature_database: str, email: str) -> EuropePMC | PubMed | None:
    """Generate an object with IDs from the specified literature database.

    Args:
        query (str): Search term for PubMed or Europe PMC.
        literature_database (str): Database to search: PubMed or Europe PMC.
        email (str): Email address to follow PubMed - NCBI guidelines.
    """
    if literature_database == "pubmed":
        Entrez.email = email
        pubmed = PubMed(query)
        ids = pubmed.get_id()
        pubmed.id_list = ids
        return pubmed
    if literature_database == "europepmc":
        europepmc = EuropePMC(query)
        ids = europepmc.get_id()
        europepmc.id_list = ids
        return europepmc
    return None


def convert_metadata_structures_to_df(
    search_code: str,
    query: str,
    literature_database: str,
    publications_metadata: list,
) -> list[list]:
    """Convert list of publication metadata structures to dataframe."""
    return [
        [
            pub.publication_id,
            pub.publication_date,
            pub.authors,
            pub.title,
            query,
            search_code,
            time.strftime("%Y-%m-%d"),
            literature_database,
        ]
        for pub in publications_metadata
    ]


def generate_publications_to_read(database_path: str, search_code: str, metadata_df: list[list]) -> None:
    """Generate publications to read based on existing database of read publications."""
    read_publications = pd.read_excel(database_path)
    existing_ids = set(read_publications["id"].dropna().astype(str))
    to_read_data = [row for row in metadata_df if str(row[0]) not in existing_ids]

    df_to_read_data = pd.DataFrame(
        to_read_data,
        columns=["id", "year_publication", "authors", "title", "search_term", "search_code", "search_date", "database"],
    )
    df_to_read_data.to_excel(f"src/aoptk/application/read_{search_code}.xlsx", index=False)


def update_master_table_search_codes(
    master_table_path: str,
    search_code: str,
    metadata_df: list[list],
) -> None:
    """Update master table with new search codes."""
    master_wb = load_workbook(master_table_path)
    master_ws = master_wb.active
    header = [cell.value for cell in master_ws[1]]
    master_id_col = get_column_index(header, "ID")
    master_search_code = get_column_index(header, "Search code")
    master_id_map = create_map_of_ids_from_master_table(master_ws, master_id_col)
    to_read_publications_id = [str(row[0]) for row in metadata_df if row[0] is not None]
    common_ids_to_read_publications_master = set(to_read_publications_id).intersection(master_id_map.keys())
    for row in metadata_df:
        row_id = str(row[0])
        if row_id in common_ids_to_read_publications_master:
            for excel_row_idx in master_id_map[row_id]:
                cell = master_ws.cell(row=excel_row_idx, column=master_search_code + 1)
                current_value = cell.value
                updated_value = f"{current_value} ; {search_code}" if current_value else search_code
                cell.value = updated_value
    master_wb.save("src/aoptk/application/updated_master_table.xlsx")


def create_map_of_ids_from_master_table(master_ws: object, master_id_col: int) -> dict[str, list[int]]:
    """Create a map of IDs from the master table to their corresponding row indices."""
    master_id_map = {}
    for idx, row in enumerate(master_ws.iter_rows(min_row=2, values_only=True), start=2):
        row_id = str(row[master_id_col]) if row[master_id_col] else None
        if row_id:
            master_id_map.setdefault(row_id, []).append(idx)
    return master_id_map


def get_column_index(header: list[str], col_name: str) -> int:
    """Get the index of a column in the header."""
    try:
        return header.index(col_name)
    except ValueError:
        sys.exit(1)
